from openai import OpenAI
import pandas as pd
from openpyxl import load_workbook
from datetime import datetime
import os
from tqdm import tqdm
import json
company = ""
industry = ""

### FIELDS TO MODIFY ###

GPT_KEY = 'your-key-here' # Your OpenAI key
CONTEXT = True # Set to True if your data has columns for "Context-Pre" and "Context-Post", False otherwise.
CATEGORIES = ["professional", "excited", "optimistic", "ambitious", "practical"] # The list of categories you wish for GPT to consider.
MAX_CATEGORIES = 3 # Maximum number of categories GPT may select
PROMPT = """Classify the category of the following sentence. Choose between "professional", "excited", "optimistic", "ambitious", or "practical".""" # Your prompt for GPT
DATA_SOURCE = "mentions.xlsx" # Your data source with sentences to be annotated
BATCH_SIZE = 1000
COLUMN_NAMES = ["Company", "Date", "Industry", "Filename"] # Any column names besides context and sentence

### FIELDS TO MODIFY ###


client = OpenAI(api_key=GPT_KEY)
OUTPUT_FILENAME = "results.xlsx" 

# Call GPT on each chunk and save results
def gpt3(data, timeout_seconds=60):

    # These are for S&P 500 specific use-case
    if "Company" in data.keys():
        company = data["Company"]
    if "Sentence" in data.keys():
        sentence = data["Sentence"]
    if "Date" in data.keys():
        date = data["Date"]
    if "Industry" in data.keys():
        industry = data["Industry"]
    context_pre = ""
    context_post = ""
    if CONTEXT:
        context_pre = data["Context-Pre"]
        context_post = data["Context-Post"]


    response = client.chat.completions.create(model="gpt-3.5-turbo-1106",
    logprobs= True,
    response_format={ "type": "json_object" },
    messages=[
        {"role": "system", "content": "You are a helpful assistant designed to output JSON."},
        {"role": "user", "content": f"""{PROMPT}
            Your output should be formatted like this: {{"categories": [{{"category": "Your category choice here", "reasoning": "Your reasoning here"}}]}}.
            Select up to {MAX_CATEGORIES} categories.
            The sentence is: 
            {context_pre} {sentence} {context_post}"""}
        ],
    timeout = timeout_seconds,
    temperature = 0.7)
    return (response.choices[0].message.content, response.choices[0].logprobs)


def extract_relevant_logprobs(logprobs_data):

    # Define the list of tokens you are interested in
    interested_tokens = CATEGORIES
    filtered_logprobs = []

    # Extract the logprobs for the specific tokens
    for i in range(len(logprobs_data.content)-1):
        logprobObject = logprobs_data.content[i]
    # for logprobObject in logprobs_data.content:

        token = logprobObject.token.strip()
        logprob = logprobObject.logprob
        if token in interested_tokens:
            filtered_logprobs.append({"token": token, "logprob": logprob})
        else:
            logprobObject_next = logprobs_data.content[i+1]
            token = token + logprobObject_next.token.strip()
            if token in interested_tokens:
                filtered_logprobs.append({"token": token, "logprob": logprob})

        

    # Convert the list of dictionaries to a JSON formatted string
    json_output = json.dumps(filtered_logprobs, indent=4)
    return json_output


def try_gpt(data, n):
    try:
        (response, logprobs) = gpt3(data)
        return (response, extract_relevant_logprobs(logprobs))
    except Exception as e:
        print("An error occurred:", type(e).__name__, ":", str(e))
        if n < 4:
            return try_gpt(data, n+1)
        else:
            print("EXCEPTION")
            return (None, None)



def copy_excel_with_timestamp():
    # Load the source workbook
    source_path = OUTPUT_FILENAME
    workbook = load_workbook(source_path)

    # Get the current timestamp
    timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")

    # Define the target workbook path with the timestamp
    target_path = f'BACKUP-OUTPUT-{timestamp}.xlsx'

    # Save the workbook as a new file with the timestamp
    workbook.save(target_path)

    print(f'Copied "{source_path}" to "{target_path}".')

def extract_names(response):
    # Check if the 'categories' key exists in the JSON object
    if 'categories' in response:
        # Extract the 'name' from each category and append to a list
        names = [category['category'] for category in response['categories']]
        return names
    else:
        return []


# Load the Excel file
df = pd.read_excel(DATA_SOURCE, engine='openpyxl')


# Check if the file exists
file_exists = os.path.isfile(OUTPUT_FILENAME)

# Determine the mode and optionally load the workbook
if file_exists:
    book = load_workbook(OUTPUT_FILENAME)
    start_row = book['Sheet1'].max_row

# Select rows START_INDEX through END_INDEX [inclusive, exclusive]
subset_df = df.iloc[start_row:start_row+BATCH_SIZE]


# Iterate through each row in the subset DataFrame
# for index, row in subset_df.iterrows():
# Initialize tqdm with the total and a more descriptive bar format if desired
for index, row in tqdm(subset_df.iterrows(), total=subset_df.shape[0], desc="Processing", unit="row"):
    # Extract the necessary information
    # company = row['Company']
    sentence = row['Sentence']
    # date = row['Date']
    # industry = row['Industry']

    for colname in COLUMN_NAMES:
        data[colname] = row[colname]
    

    data = {}
    # data["Filename"] = row['Filename']
    # data["Company"] = company
    data["Sentence"] = sentence
    # data["Date"] = date
    # data["Industry"] = industry

    if CONTEXT:
        context_pre = row['Context-Pre']
        context_post = row['Context-Post']
        data["Context-Pre"] = context_pre
        data["Context-Post"] = context_post


    (response, logprobs) = try_gpt(data, 1)

    # print(type(response))  # Check the type of response
    # print(response)        # Print the content of response

    response_dict = json.loads(response)

    data["Output"] = response
    data["LogProbs"] = logprobs
    categories = extract_names(response_dict)

    # Add categories as individual columns to the data dictionary
    for i in range(0, MAX_CATEGORIES):
        data[f'Category {i+1}'] = ""

    # Add categories as individual columns to the data dictionary
    for i, category in enumerate(categories):
        data[f'Category {i+1}'] = [category]
    

    # Create a new DataFrame for the current row's API response
    new_row_df = pd.DataFrame(data)

    # Check if the file exists
    file_exists = os.path.isfile(OUTPUT_FILENAME)

    # Determine the mode and optionally load the workbook
    if file_exists:
        book = load_workbook(OUTPUT_FILENAME)
        with pd.ExcelWriter(OUTPUT_FILENAME, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
            start_row = book['Sheet1'].max_row
            new_row_df.to_excel(writer, startrow=start_row, index=False, header=False)

    else:
        with pd.ExcelWriter(OUTPUT_FILENAME, engine='openpyxl', mode='w') as writer:
            book = None
            new_row_df.to_excel(writer, index=False)


copy_excel_with_timestamp()


